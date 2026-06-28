#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_data.py — בונה את data.json שהאתר (מאגר רופאים בהסדר) קורא ממנו.

מקורות:
  • מנורה — CSV ציבורי (UTF-8, 14 עמודות).
  • כלל   — CSV ציבורי (Windows-1255, 13 עמודות).
  • איילון — CSV ציבורי (UTF-8, 15 עמודות, כולל בית חולים).
  • שאר החברות (מגדל/פניקס) — מתוך ה-PAYLOAD המוטמע ב-index.html (קפוא).

הרצה:
  python build_data.py --html index.html --out data.json \
      [--menora menora.csv] [--clal clal.csv] [--ayalon ayalon.csv]

בדיקות שפיות: אם מספר רשומות של חברה יוצא מחוץ לטווח הצפוי — הסקריפט נכשל
(exit 1) ולא דורס את data.json.
"""

import argparse, csv, json, re, sys
from datetime import datetime

SANITY = {"מנורה": (1500, 6000), "כלל": (1000, 4000), "איילון": (800, 3000), "פניקס": (800, 3000), "הראל": (800, 3000)}

TITLE_MAP = {
    "דר": "ד\"ר", "דר'": "ד\"ר", "ד\"ר": "ד\"ר", "ד״ר": "ד\"ר", "דוקטור": "ד\"ר",
    "פרופ": "פרופ'", "פרופ'": "פרופ'", "פרופ׳": "פרופ'", "פרופסור": "פרופ'",
    "מר": "מר", "גב": "גב", "גב'": "גב", "גברת": "גב",
}


def norm_title(t):
    t = (t or "").strip().rstrip(".").strip()
    return TITLE_MAP.get(t, t)


def yn(v):
    # מנורה/כלל: כן/לא  |  איילון: כ/ל
    return 1 if str(v).strip() in ("כן", "כ") else 0


def is_expired(end):
    end = (end or "").strip()
    if not end:
        return False
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(end, fmt).date() < datetime.now().date()
        except ValueError:
            continue
    return False


def extract_payload(html_path):
    with open(html_path, "r", encoding="utf-8") as f:
        html = f.read()
    m = re.search(r"const\s+PAYLOAD\s*=\s*", html)
    if not m:
        sys.exit("ERROR: לא נמצא PAYLOAD ב-HTML")
    i = html.index("{", m.end())
    depth, j, in_str, esc = 0, i, False, False
    while j < len(html):
        ch = html[j]
        if in_str:
            if esc: esc = False
            elif ch == "\\": esc = True
            elif ch == '"': in_str = False
        else:
            if ch == '"': in_str = True
            elif ch == "{": depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    return json.loads(html[i:j + 1])
        j += 1
    sys.exit("ERROR: PAYLOAD לא תקין")


def decode_payload(p):
    comp, tit, spec, subs, cities, rows = (
        p["companies"], p["titles"], p["specialties"], p["subs"], p["cities"], p["rows"])
    out = []
    for i, r in enumerate(rows):
        out.append({
            "id": "doc_" + str(i), "company": comp[r[0]], "name": r[1] or "—",
            "lic": r[2] or "", "title": tit[r[3]] or "", "specialty": spec[r[4]] or "",
            "sub": subs[r[5]] or "", "s3": r[6], "s1": r[7],
            "cities": [cities[ci] for ci in (r[8] or [])], "hospital": r[9] or "",
            "start": r[10] or "", "end": r[11] or "", "expired": bool(r[12]),
            "lkey": r[13] or "",
        })
    return out


def parse_menora(csv_path):
    """מנורה: UTF-8, 14 עמודות, עד 4 ערים, התמחות נוספת אחת, ללא בית חולים."""
    recs = []
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f); next(reader, None)
        for n, row in enumerate(reader):
            if not row or len(row) < 14:
                continue
            (first, last, lic, title, spec, sub,
             s3, s1, c1, c2, c3, c4, start, end) = [c.strip() for c in row[:14]]
            recs.append({
                "id": "menora_" + str(n), "company": "מנורה",
                "name": (first + " " + last).strip() or "—", "lic": lic,
                "title": norm_title(title), "specialty": spec, "sub": sub,
                "s3": yn(s3), "s1": yn(s1),
                "cities": [c for c in (c1, c2, c3, c4) if c], "hospital": "",
                "start": start, "end": end, "expired": is_expired(end),
                "lkey": re.sub(r"\D", "", lic),
            })
    return recs


def parse_clal(csv_path):
    """כלל: Windows-1255, 13 עמודות, 2 ערים, 2 עמודות התמחות נוספת, ללא בית חולים."""
    recs = []
    with open(csv_path, "r", encoding="cp1255", newline="") as f:
        reader = csv.reader(f); next(reader, None)
        for n, row in enumerate(reader):
            if not row or len(row) < 13:
                continue
            (first, last, lic, title, spec, sub1,
             s3, s1, c1, c2, sub2, start, end) = [c.strip() for c in row[:13]]
            sub = ", ".join([s for s in (sub1, sub2) if s])
            recs.append({
                "id": "clal_" + str(n), "company": "כלל",
                "name": (first + " " + last).strip() or "—", "lic": lic,
                "title": norm_title(title), "specialty": spec, "sub": sub,
                "s3": yn(s3), "s1": yn(s1),
                "cities": [c for c in (c1, c2) if c], "hospital": "",
                "start": start, "end": end, "expired": is_expired(end),
                "lkey": re.sub(r"\D", "", lic),
            })
    return recs


def parse_ayalon(csv_path):
    """איילון: UTF-8, 15 עמודות, 4 ערים, כולל בית חולים. ניתוחים=כ/ל, סדר הפוך (s1 לפני s3)."""
    recs = []
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f); next(reader, None)
        for n, row in enumerate(reader):
            if not row or len(row) < 15:
                continue
            (first, last, lic, title, spec, sub,
             c1, c2, c3, c4, hosp, s1, s3, start, end) = [c.strip() for c in row[:15]]
            recs.append({
                "id": "ayalon_" + str(n), "company": "איילון",
                "name": (first + " " + last).strip() or "—", "lic": lic,
                "title": norm_title(title), "specialty": spec, "sub": sub,
                "s3": yn(s3), "s1": yn(s1),
                "cities": [c for c in (c1, c2, c3, c4) if c], "hospital": hosp,
                "start": start, "end": end, "expired": is_expired(end),
                "lkey": re.sub(r"\D", "", lic),
            })
    return recs


def parse_phoenix(csv_path):
    """פניקס: UTF-8, 14 עמודות. שם מלא בשדה אחד, 2 זוגות תאריכים, sub מופרד ב-';', 3 ערים, ללא בית חולים."""
    recs = []
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f); next(reader, None)
        for n, row in enumerate(reader):
            if not row or len(row) < 14:
                continue
            (name, lic, title, spec, sub, start1, end1, start2, end2,
             surg3, surg1, c1, c2, c3) = [c.strip() for c in row[:14]]
            # שני טווחי התקשרות אפשריים — אם יש טווח פתוח (בלי תאריך סיום) הרופא פעיל
            ranges = [(start1, end1), (start2, end2)]
            starts = [s for s, e in ranges if s]
            ends = [e for s, e in ranges if e]
            has_open = any(s and not e for s, e in ranges)
            start = starts[0] if starts else ""
            end = "" if has_open else (ends[-1] if ends else "")
            sub_clean = ", ".join([p.strip() for p in sub.split(";") if p.strip()])
            recs.append({
                "id": "phoenix_" + str(n), "company": "פניקס",
                "name": name or "—", "lic": lic,
                "title": norm_title(title), "specialty": spec, "sub": sub_clean,
                "s3": yn(surg3), "s1": yn(surg1),
                "cities": [c for c in (c1, c2, c3) if c], "hospital": "",
                "start": start, "end": end, "expired": is_expired(end),
                "lkey": re.sub(r"\D", "", lic),
            })
    return recs


def parse_harel(xlsx_path):
    """הראל: קובץ Excel (xlsx), 18 עמודות, עם עמודת 'רופא פעיל' מפורשת (כן/לא)."""
    import openpyxl  # נדרש רק אם משתמשים ב-Harel
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    next(rows, None)  # שורת כותרות
    recs = []
    for n, row in enumerate(rows):
        if not row or len(row) < 17 or not (row[0] or row[1]):
            continue

        def g(i):
            v = row[i] if i < len(row) else None
            return "" if v is None else str(v).strip()

        lic = g(2)
        sub = ", ".join([p.strip() for p in (g(5) + " ; " + g(6)).split(";") if p.strip()])
        created = row[15] if len(row) > 15 else None
        start = created.strftime("%d/%m/%Y") if hasattr(created, "strftime") else g(15)
        recs.append({
            "id": "harel_" + str(n), "company": "הראל",
            "name": (g(0) + " " + g(1)).strip() or "—", "lic": lic,
            "title": norm_title(g(3)), "specialty": g(4), "sub": sub,
            "s3": yn(g(8)), "s1": yn(g(7)),
            "cities": [g(9)] if g(9) else [], "hospital": "",
            "start": start, "end": "",
            "expired": (g(16) != "כן"),   # 'רופא פעיל' = כן/לא
            "lkey": re.sub(r"\D", "", lic),
        })
    return recs


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--html", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--menora")
    ap.add_argument("--clal")
    ap.add_argument("--ayalon")
    ap.add_argument("--phoenix")
    ap.add_argument("--harel")
    args = ap.parse_args()

    fresh, provided = [], set()
    for company, path, parser in (
        ("מנורה", args.menora, parse_menora),
        ("כלל", args.clal, parse_clal),
        ("איילון", args.ayalon, parse_ayalon),
        ("פניקס", args.phoenix, parse_phoenix),
        ("הראל", args.harel, parse_harel),
    ):
        if not path:
            continue
        recs = parser(path)
        lo, hi = SANITY.get(company, (1, 10 ** 9))
        if not (lo <= len(recs) <= hi):
            sys.exit(f"ERROR: {company} — {len(recs)} רשומות, מחוץ לטווח [{lo}-{hi}]. "
                     f"לא כותב data.json.")
        fresh += recs
        provided.add(company)

    others = [d for d in decode_payload(extract_payload(args.html))
              if d["company"] not in provided]
    data = others + fresh

    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))

    by = {}
    for d in data:
        by[d["company"]] = by.get(d["company"], 0) + 1
    print(f"OK  data.json: {len(data)} רשומות")
    for c, n in sorted(by.items(), key=lambda x: -x[1]):
        tag = "  (מעודכן)" if c in provided else "  (קפוא)"
        print(f"     {c}: {n}{tag}")


if __name__ == "__main__":
    main()
